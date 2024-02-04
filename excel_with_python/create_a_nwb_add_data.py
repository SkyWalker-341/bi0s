from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

data = ['Name ', 'Roll no', 'phone number', 'aneesh', 23001, '*******', 'abin', 23002, '*******', 'sam', 23003, '*******', 'arvind', 23004, '*******', 'deepak', 23005, '*******', 'sai ', 23006, 
'********', 'richu ', 23007, '********']
def creates_wb():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["name","age","stutes","number_c"])
    wb.save("simple.xlsx")

def loading_wb():
    wb = load_workbook("""E:/python/Bi0s/excel_with_python/simple.xlsx""")
    ws= wb.active
    for i in range(1,15):
        for j in range(1,5):
            char = get_column_letter(j)
            result = ws[char +str(i)] = char +str(i)
            print(result)
        wb.save("""E:/python/Bi0s/excel_with_python/simple.xlsx""")

print("welocme to excel with python"  + "press 1 for creat_file " +" OR "+ "press 2 for loading the file")
choise = int(input("enteh-r the you  chiose :: ")) 
if choise == 1:
    creates_wb()
else:
    loading_wb()