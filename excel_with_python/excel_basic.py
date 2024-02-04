from openpyxl import workbook, load_workbook
'''
wb = load_workbook("""E:/python/Bi0s/excel_with_python/test_1.xlsx""")
print(wb)
ws = wb.active
#for i in range(2,9):
#print(ws["A"+str(i)].value) 
change = (ws["A2"].value) = "aneesh"
wb.save("""E:/python/Bi0s/excel_with_python/test_1.xlsx""")
#print()'''



wb = load_workbook("""E:/python/Bi0s/excel_with_python/test_1.xlsx""")
print(wb)
ws = wb.active
m_row = ws.max_row
m_column = ws.max_column
result = []
for i in range(1,m_row+1):
    for j in range(1,m_column+1):
            cell = ws.cell(row = i ,column = j).value
            result.append(cell)
print(result)
wb.save("""E:/python/Bi0s/excel_with_python/test_1.xlsx""")
#print()
