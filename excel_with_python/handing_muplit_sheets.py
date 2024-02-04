from openpyxl import workbook, load_workbook
'''
wb = load_workbook("""E:/python/Bi0s/excel_with_python/test_1.xlsx""")
print(wb.sheetnames)

'''
'''
wb = load_workbook("""E:/python/Bi0s/excel_with_python/test_1.xlsx""")
ws =wb["Sheet2"]                                                            # to select the sheets
print(wb["Sheet2"])


wb = load_workbook("""E:/python/Bi0s/excel_with_python/test_1.xlsx""")
wb.create_sheet("simple")                                                   # to create a sheets 
print(wb.sheetnames)

                                                                                           
'''